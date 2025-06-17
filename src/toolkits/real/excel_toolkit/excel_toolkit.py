from langchain_core.tools import tool
from mcp.server.fastmcp import FastMCP
import openpyxl
import os
from typing import Optional


mcp = FastMCP('excel_toolkit')
def excel_create_workbook(file_path: str) -> dict:
    """
    Create a new Excel workbook and save it to the specified file path.

    Args:
        file_path (str): The path to save the Excel workbook.

    Returns:
        dict: A dictionary containing the success status and message.
    """
    dir_name = os.path.dirname(file_path)
    if not os.path.exists(dir_name):
        os.makedirs(dir_name, exist_ok=True)
    workbook = openpyxl.Workbook()
    workbook.save(file_path)
    return {'success': True, 'message': f'Workbook created at {file_path}'}


@tool(parse_docstring=True)
def langgraph_excel_create_workbook(file_path: str) -> dict:
    """
    Create a new Excel workbook and save it to the specified file path.

    Args:
        file_path (str): The path to save the Excel workbook.

    Returns:
        dict: A dictionary containing the success status and message.
    """
    return excel_create_workbook(file_path=file_path)


@mcp.tool()
def mcp_excel_create_workbook(file_path: str) -> dict:
    """
    Create a new Excel workbook and save it to the specified file path.

    Args:
        file_path (str): The path to save the Excel workbook.

    Returns:
        dict: A dictionary containing the success status and message.
    """
    return excel_create_workbook(file_path=file_path)


def excel_add_sheet(file_path: str, sheet_name: str) -> dict:
    """
    Add a new sheet to an existing Excel workbook.

    Args:
        file_path (str): The path to the Excel workbook.
        sheet_name (str): The name of the new sheet.

    Returns:
        dict: A dictionary containing the success status and message.
    """
    if not os.path.exists(file_path):
        return {'success': False, 'message': f'File {file_path} does not exist.'}
    workbook = openpyxl.load_workbook(file_path)
    workbook.create_sheet(title=sheet_name)
    workbook.save(file_path)
    return {'success': True, 'message': f"Sheet '{sheet_name}' added to {file_path}"}


@tool(parse_docstring=True)
def langgraph_excel_add_sheet(file_path: str, sheet_name: str) -> dict:
    """
    Add a new sheet to an existing Excel workbook.

    Args:
        file_path (str): The path to the Excel workbook.
        sheet_name (str): The name of the new sheet.

    Returns:
        dict: A dictionary containing the success status and message.
    """
    return excel_add_sheet(file_path=file_path, sheet_name=sheet_name)


@mcp.tool()
def mcp_excel_add_sheet(file_path: str, sheet_name: str) -> dict:
    """
    Add a new sheet to an existing Excel workbook.

    Args:
        file_path (str): The path to the Excel workbook.
        sheet_name (str): The name of the new sheet.

    Returns:
        dict: A dictionary containing the success status and message.
    """
    return excel_add_sheet(file_path=file_path, sheet_name=sheet_name)


def excel_delete_sheet(file_path: str, sheet_name: str) -> dict:
    """
    Delete a sheet from an existing Excel workbook.

    Args:
        file_path (str): The path to the Excel workbook.
        sheet_name (str): The name of the sheet to delete.

    Returns:
        dict: A dictionary containing the success status and message.
    """
    if not os.path.exists(file_path):
        return {'success': False, 'message': f'File {file_path} does not exist.'}
    workbook = openpyxl.load_workbook(file_path)
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
        workbook.save(file_path)
        return {'success': True, 'message': f"Sheet '{sheet_name}' deleted from {file_path}"}
    else:
        return {'success': False, 'message': f"Sheet '{sheet_name}' does not exist in {file_path}"}


@tool(parse_docstring=True)
def langgraph_excel_delete_sheet(file_path: str, sheet_name: str) -> dict:
    """
    Delete a sheet from an existing Excel workbook.

    Args:
        file_path (str): The path to the Excel workbook.
        sheet_name (str): The name of the sheet to delete.

    Returns:
        dict: A dictionary containing the success status and message.
    """
    return excel_delete_sheet(file_path=file_path, sheet_name=sheet_name)


@mcp.tool()
def mcp_excel_delete_sheet(file_path: str, sheet_name: str) -> dict:
    """
    Delete a sheet from an existing Excel workbook.

    Args:
        file_path (str): The path to the Excel workbook.
        sheet_name (str): The name of the sheet to delete.

    Returns:
        dict: A dictionary containing the success status and message.
    """
    return excel_delete_sheet(file_path=file_path, sheet_name=sheet_name)


def excel_write_cell(file_path: str, cell: str, value: str, sheet_name: Optional[str]='Sheet') -> dict:
    """
    Write a value to a specific cell in an existing Excel workbook.

    Args:
        file_path (str): The path to the Excel workbook.
        cell (str): The cell to write to (e.g., "A1").
        value (str): The value to write.
        sheet_name (str): The name of the sheet.

    Returns:
        dict: A dictionary containing the success status and message.
    """
    if not os.path.exists(file_path):
        return {'success': False, 'message': f'File {file_path} does not exist.'}
    workbook = openpyxl.load_workbook(file_path)
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheet[cell] = value
        workbook.save(file_path)
        return {'success': True, 'message': f"Value '{value}' written to cell '{cell}' in sheet '{sheet_name}' of {file_path}"}
    else:
        return {'success': False, 'message': f"Sheet '{sheet_name}' does not exist in {file_path}"}


@tool(parse_docstring=True)
def langgraph_excel_write_cell(file_path: str, cell: str, value: str, sheet_name: Optional[str]='Sheet') -> dict:
    """
    Write a value to a specific cell in an existing Excel workbook.

    Args:
        file_path (str): The path to the Excel workbook.
        cell (str): The cell to write to (e.g., "A1").
        value (str): The value to write.
        sheet_name (str): The name of the sheet.

    Returns:
        dict: A dictionary containing the success status and message.
    """
    return excel_write_cell(file_path=file_path, cell=cell, value=value, sheet_name=sheet_name)


@mcp.tool()
def mcp_excel_write_cell(file_path: str, cell: str, value: str, sheet_name: Optional[str]='Sheet') -> dict:
    """
    Write a value to a specific cell in an existing Excel workbook.

    Args:
        file_path (str): The path to the Excel workbook.
        cell (str): The cell to write to (e.g., "A1").
        value (str): The value to write.
        sheet_name (str): The name of the sheet.

    Returns:
        dict: A dictionary containing the success status and message.
    """
    return excel_write_cell(file_path=file_path, cell=cell, value=value, sheet_name=sheet_name)


def excel_read_cell(file_path: str, cell: str, sheet_name: Optional[str]='Sheet') -> dict:
    """
    Read a value from a specific cell in an existing Excel workbook.

    Args:
        file_path (str): The path to the Excel workbook.
        cell (str): The cell to read from (e.g., "A1").
        sheet_name (str): The name of the sheet.

    Returns:
        dict: A dictionary containing the success status and the value read from the cell.
    """
    if not os.path.exists(file_path):
        return {'success': False, 'message': f'File {file_path} does not exist.'}
    workbook = openpyxl.load_workbook(file_path)
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        value = sheet[cell].value
        return {'success': True, 'value': value}
    else:
        return {'success': False, 'message': f"Sheet '{sheet_name}' does not exist in {file_path}"}


@tool(parse_docstring=True)
def langgraph_excel_read_cell(file_path: str, cell: str, sheet_name: Optional[str]='Sheet') -> dict:
    """
    Read a value from a specific cell in an existing Excel workbook.

    Args:
        file_path (str): The path to the Excel workbook.
        cell (str): The cell to read from (e.g., "A1").
        sheet_name (str): The name of the sheet.

    Returns:
        dict: A dictionary containing the success status and the value read from the cell.
    """
    return excel_read_cell(file_path=file_path, cell=cell, sheet_name=sheet_name)


@mcp.tool()
def mcp_excel_read_cell(file_path: str, cell: str, sheet_name: Optional[str]='Sheet') -> dict:
    """
    Read a value from a specific cell in an existing Excel workbook.

    Args:
        file_path (str): The path to the Excel workbook.
        cell (str): The cell to read from (e.g., "A1").
        sheet_name (str): The name of the sheet.

    Returns:
        dict: A dictionary containing the success status and the value read from the cell.
    """
    return excel_read_cell(file_path=file_path, cell=cell, sheet_name=sheet_name)


def excel_format_cell(file_path: str, cell: str, format: str, sheet_name: Optional[str]='Sheet') -> dict:
    """
    Format a cell in an existing Excel workbook.

    Args:
        file_path (str): The path to the Excel workbook.
        cell (str): The cell to format (e.g., "A1").
        format (str): The format to apply (e.g., "0.00").
        sheet_name (str): The name of the sheet.

    Returns:
        dict: A dictionary containing the success status and message.
    """
    if not os.path.exists(file_path):
        return {'success': False, 'message': f'File {file_path} does not exist.'}
    workbook = openpyxl.load_workbook(file_path)
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheet[cell].number_format = format
        workbook.save(file_path)
        return {'success': True, 'message': f"Cell '{cell}' formatted with '{format}' in sheet '{sheet_name}' of {file_path}"}
    else:
        return {'success': False, 'message': f"Sheet '{sheet_name}' does not exist in {file_path}"}


@tool(parse_docstring=True)
def langgraph_excel_format_cell(file_path: str, cell: str, format: str, sheet_name: Optional[str]='Sheet') -> dict:
    """
    Format a cell in an existing Excel workbook.

    Args:
        file_path (str): The path to the Excel workbook.
        cell (str): The cell to format (e.g., "A1").
        format (str): The format to apply (e.g., "0.00").
        sheet_name (str): The name of the sheet.

    Returns:
        dict: A dictionary containing the success status and message.
    """
    return excel_format_cell(file_path=file_path, cell=cell, format=format, sheet_name=sheet_name)


@mcp.tool()
def mcp_excel_format_cell(file_path: str, cell: str, format: str, sheet_name: Optional[str]='Sheet') -> dict:
    """
    Format a cell in an existing Excel workbook.

    Args:
        file_path (str): The path to the Excel workbook.
        cell (str): The cell to format (e.g., "A1").
        format (str): The format to apply (e.g., "0.00").
        sheet_name (str): The name of the sheet.

    Returns:
        dict: A dictionary containing the success status and message.
    """
    return excel_format_cell(file_path=file_path, cell=cell, format=format, sheet_name=sheet_name)


def excel_read_sheet(file_path: str, sheet_name: Optional[str]='Sheet') -> dict:
    """
    Read the content of a specific sheet in an existing Excel workbook.

    Args:
        file_path (str): The path to the Excel workbook.
        sheet_name (str): The name of the sheet.

    Returns:
        dict: A dictionary containing the success status and the content of the sheet.
    """
    if not os.path.exists(file_path):
        return {'success': False, 'message': f'File {file_path} does not exist.'}
    workbook = openpyxl.load_workbook(file_path)
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        return {'success': True, 'data': data}
    else:
        return {'success': False, 'message': f"Sheet '{sheet_name}' does not exist in {file_path}"}


@tool(parse_docstring=True)
def langgraph_excel_read_sheet(file_path: str, sheet_name: Optional[str]='Sheet') -> dict:
    """
    Read the content of a specific sheet in an existing Excel workbook.

    Args:
        file_path (str): The path to the Excel workbook.
        sheet_name (str): The name of the sheet.

    Returns:
        dict: A dictionary containing the success status and the content of the sheet.
    """
    return excel_read_sheet(file_path=file_path, sheet_name=sheet_name)


@mcp.tool()
def mcp_excel_read_sheet(file_path: str, sheet_name: Optional[str]='Sheet') -> dict:
    """
    Read the content of a specific sheet in an existing Excel workbook.

    Args:
        file_path (str): The path to the Excel workbook.
        sheet_name (str): The name of the sheet.

    Returns:
        dict: A dictionary containing the success status and the content of the sheet.
    """
    return excel_read_sheet(file_path=file_path, sheet_name=sheet_name)


def excel_list_sheets(file_path: str) -> dict:
    """
    List all sheets in an existing Excel workbook.

    Args:
        file_path (str): The path to the Excel workbook.

    Returns:
        dict: A dictionary containing the success status and the list of sheet names.
    """
    if not os.path.exists(file_path):
        return {'success': False, 'message': f'File {file_path} does not exist.'}
    workbook = openpyxl.load_workbook(file_path)
    sheet_names = workbook.sheetnames
    return {'success': True, 'sheets': sheet_names}


@tool(parse_docstring=True)
def langgraph_excel_list_sheets(file_path: str) -> dict:
    """
    List all sheets in an existing Excel workbook.

    Args:
        file_path (str): The path to the Excel workbook.

    Returns:
        dict: A dictionary containing the success status and the list of sheet names.
    """
    return excel_list_sheets(file_path=file_path)


@mcp.tool()
def mcp_excel_list_sheets(file_path: str) -> dict:
    """
    List all sheets in an existing Excel workbook.

    Args:
        file_path (str): The path to the Excel workbook.

    Returns:
        dict: A dictionary containing the success status and the list of sheet names.
    """
    return excel_list_sheets(file_path=file_path)


